# -*- coding: utf-8 -*-
"""
W2O SALADA 새벽배송 손익분기 시뮬레이터 생성 스크립트
- 주 2회(화·목) 새벽배송 기준
- 가정값 시트의 셀만 바꾸면 손익/손익분기가 자동 재계산되는 수식 기반 모델
- 가정값은 한국 새벽배송 업계 일반 추정치로 초기화
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WON = '#,##0"원"'
WON_PLAIN = '#,##0'
PCT = '0.0%'
NUM1 = '0.0'

# ---- 스타일 ----
title_font = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1D9E75")
sec_font = Font(name="맑은 고딕", size=11, bold=True, color="0A1A0F")
sec_fill = PatternFill("solid", fgColor="C8EDDD")
label_font = Font(name="맑은 고딕", size=10)
input_font = Font(name="맑은 고딕", size=10, bold=True, color="9C5700")
input_fill = PatternFill("solid", fgColor="FFF3C4")  # 노란색 = 입력 셀
calc_font = Font(name="맑은 고딕", size=10)
result_font = Font(name="맑은 고딕", size=11, bold=True)
result_fill = PatternFill("solid", fgColor="EAF6F1")
note_font = Font(name="맑은 고딕", size=9, italic=True, color="6B7280")

thin = Side(style="thin", color="D0D7DE")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_block(ws, cell_range, font=None, fill=None, fmt=None, border=False, align=None):
    for row in ws[cell_range]:
        for c in row:
            if font:
                c.font = font
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            if border:
                c.border = box
            if align:
                c.alignment = align


wb = Workbook()

# =========================================================
# 1) 가정값 시트
# =========================================================
A = wb.active
A.title = "가정값"
A.sheet_view.showGridLines = False
A.column_dimensions["A"].width = 3
A.column_dimensions["B"].width = 26
A.column_dimensions["C"].width = 16
A.column_dimensions["D"].width = 10
A.column_dimensions["E"].width = 44

A["B1"] = "W2O SALADA · 새벽배송 손익분기 시뮬레이터"
A.merge_cells("B1:E1")
A["B1"].font = title_font
A["B1"].fill = title_fill
A["B1"].alignment = Alignment(horizontal="center", vertical="center")
A.row_dimensions[1].height = 30

A["B2"] = "노란 셀 = 직접 입력하는 가정값 · 나머지 시트는 자동 계산됩니다 (주 2회 화·목 배송 기준)"
A.merge_cells("B2:E2")
A["B2"].font = note_font

# (라벨, 값, 단위, 셀주소키, 설명)
ref = {}
row = 4


def section(name):
    global row
    A.cell(row=row, column=2, value=name)
    A.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    style_block(A, f"B{row}:E{row}", font=sec_font, fill=sec_fill)
    A.row_dimensions[row].height = 22
    row += 1


def assume(key, label, value, unit, note, fmt=WON_PLAIN):
    global row
    A.cell(row=row, column=2, value=label).font = label_font
    vc = A.cell(row=row, column=3, value=value)
    vc.font = input_font
    vc.fill = input_fill
    vc.number_format = fmt
    vc.border = box
    vc.alignment = Alignment(horizontal="right")
    A.cell(row=row, column=4, value=unit).font = note_font
    A.cell(row=row, column=5, value=note).font = note_font
    ref[key] = f"가정값!$C${row}"
    row += 1


section("① 매출 가정")
assume("aov", "객단가 (평균 결제액)", 18000, "원", "본품 최소 11,000원 + 음료·추가구성 평균. 업계 샐러드 새벽배송 1.5~2.5만")
assume("daily", "배송일당 주문 수", 80, "건", "이 값을 바꿔가며 손익분기 확인 (기본 시나리오)")
assume("freq", "주 배송 횟수", 2, "회", "화·목 = 2회 (Weekly 2 Order)", WON_PLAIN)
assume("weeks", "월 평균 주 수", 4.345, "주", "52주 ÷ 12개월 = 4.345", NUM1)

section("② 변동비 가정 (주문 1건당)")
assume("cogs", "식자재 원가율", 0.38, "%", "샐러드·신선 HMR 통상 35~42%", PCT)
assume("waste", "폐기율 (식자재 대비)", 0.05, "%", "예측생산 신선식품 통상 3~8%", PCT)
assume("pack", "건당 포장재비", 1200, "원", "보냉박스+아이스팩+용기 (비회수 기준)")
assume("ship", "건당 위탁 배송비", 3500, "원", "새벽배송 3PL 위탁 단가 통상 3,000~4,500")
assume("pg", "PG 수수료율", 0.033, "%", "토스페이먼츠 약 3.3%", PCT)
assume("alim", "건당 알림톡비", 40, "원", "주문완료+출발+완료 3건 × 12~15원")

section("③ 고정비 가정 (월)")
assume("rent", "주방 임대·관리비", 2500000, "원", "공유주방~소형 센터 초기 기준")
assume("labor", "인건비", 12000000, "원", "조리·포장·CS 4~5명 초기 소규모")
assume("car", "차량·냉장설비", 1500000, "원", "리스·유지·냉장 인프라")
assume("mkt", "마케팅비", 3000000, "원", "신규 유입 광고 (초기 공격 시 ↑)")
assume("etc", "기타 운영비", 1000000, "원", "SW·결제·보험·잡비")

last_assume_row = row - 1

# =========================================================
# 2) 손익 시트 (월 손익계산서)
# =========================================================
P = wb.create_sheet("손익")
P.sheet_view.showGridLines = False
P.column_dimensions["A"].width = 3
P.column_dimensions["B"].width = 28
P.column_dimensions["C"].width = 16
P.column_dimensions["D"].width = 14
P.column_dimensions["E"].width = 40

P["B1"] = "월간 손익계산서 (현재 가정값 기준)"
P.merge_cells("B1:E1")
P["B1"].font = title_font
P["B1"].fill = title_fill
P["B1"].alignment = Alignment(horizontal="center", vertical="center")
P.row_dimensions[1].height = 28

pr = 3


def pline(label, formula, fmt=WON, bold=False, fill=False, note=""):
    global pr
    P.cell(row=pr, column=2, value=label).font = result_font if bold else label_font
    vc = P.cell(row=pr, column=3, value=formula)
    vc.font = result_font if bold else calc_font
    vc.number_format = fmt
    vc.border = box
    vc.alignment = Alignment(horizontal="right")
    if fill:
        P.cell(row=pr, column=2).fill = result_fill
        vc.fill = result_fill
    if note:
        P.cell(row=pr, column=5, value=note).font = note_font
    cur = f"손익!$C${pr}"
    pr += 1
    return cur


def psection(name):
    global pr
    P.cell(row=pr, column=2, value=name)
    P.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=5)
    style_block(P, f"B{pr}:E{pr}", font=sec_font, fill=sec_fill)
    pr += 1


psection("물량")
mdays = pline("월 배송일 수", f"={ref['freq']}*{ref['weeks']}", NUM1, note="주 배송횟수 × 월평균주수")
morders = pline("월 주문 건수", f"={ref['daily']}*{mdays}", WON_PLAIN, note="배송일당 주문수 × 월 배송일수")

psection("매출")
rev = pline("월 매출액", f"={ref['aov']}*{morders}", WON, bold=True, fill=True, note="객단가 × 월 주문건수")

psection("변동비 (주문 1건당)")
v_cogs = pline("식자재비", f"={ref['aov']}*{ref['cogs']}", WON, note="객단가 × 원가율")
v_waste = pline("폐기 손실", f"={v_cogs}*{ref['waste']}", WON, note="식자재비 × 폐기율")
v_pack = pline("포장재비", f"={ref['pack']}", WON)
v_ship = pline("위탁 배송비", f"={ref['ship']}", WON)
v_pg = pline("PG 수수료", f"={ref['aov']}*{ref['pg']}", WON, note="객단가 × PG율")
v_alim = pline("알림톡비", f"={ref['alim']}", WON)
v_sum = pline("건당 변동비 합계", f"={v_cogs}+{v_waste}+{v_pack}+{v_ship}+{v_pg}+{v_alim}",
              WON, bold=True, fill=True)

psection("공헌이익")
cm_unit = pline("건당 공헌이익", f"={rev}/{morders}-{v_sum}", WON, bold=True, fill=True,
                note="객단가 − 건당 변동비")
cm_rate = pline("공헌이익률", f"={cm_unit}/{ref['aov']}", PCT, bold=True)
cm_month = pline("월 공헌이익", f"={cm_unit}*{morders}", WON, bold=True, fill=True)

psection("고정비 (월)")
f_rent = pline("주방 임대·관리비", f"={ref['rent']}")
f_labor = pline("인건비", f"={ref['labor']}")
f_car = pline("차량·냉장설비", f"={ref['car']}")
f_mkt = pline("마케팅비", f"={ref['mkt']}")
f_etc = pline("기타 운영비", f"={ref['etc']}")
f_sum = pline("월 고정비 합계", f"={f_rent}+{f_labor}+{f_car}+{f_mkt}+{f_etc}",
              WON, bold=True, fill=True)

psection("영업손익")
op = pline("월 영업이익", f"={cm_month}-{f_sum}", WON, bold=True, fill=True,
           note="월 공헌이익 − 월 고정비")
op_rate = pline("영업이익률", f"=IF({rev}=0,0,{op}/{rev})", PCT, bold=True)

# =========================================================
# 3) 손익분기 시트
# =========================================================
B = wb.create_sheet("손익분기")
B.sheet_view.showGridLines = False
B.column_dimensions["A"].width = 3
B.column_dimensions["B"].width = 26
B.column_dimensions["C"].width = 16
B.column_dimensions["D"].width = 4
for col in "EFGHI":
    B.column_dimensions[col].width = 15

B["B1"] = "손익분기점 (BEP) 분석"
B.merge_cells("B1:I1")
B["B1"].font = title_font
B["B1"].fill = title_fill
B["B1"].alignment = Alignment(horizontal="center", vertical="center")
B.row_dimensions[1].height = 28

B["B3"] = "핵심 지표"
B.merge_cells("B3:C3")
style_block(B, "B3:C3", font=sec_font, fill=sec_fill)


def bline(r, label, formula, fmt=WON, bold=True):
    B.cell(row=r, column=2, value=label).font = result_font if bold else label_font
    vc = B.cell(row=r, column=3, value=formula)
    vc.font = result_font if bold else calc_font
    vc.number_format = fmt
    vc.border = box
    vc.fill = result_fill
    vc.alignment = Alignment(horizontal="right")
    B.cell(row=r, column=2).fill = result_fill
    return f"손익분기!$C${r}"


bep_orders = bline(4, "손익분기 월 주문건수", f"={f_sum}/{cm_unit}", WON_PLAIN)
bep_daily = bline(5, "손익분기 배송일당 주문수", f"={bep_orders}/{mdays}", WON_PLAIN)
bep_rev = bline(6, "손익분기 월 매출액", f"={bep_orders}*{ref['aov']}", WON)
bep_safety = bline(7, "현재 가정 대비 안전마진",
                   f"=IF({morders}=0,0,({morders}-{bep_orders})/{morders})", PCT)
bep_status = bline(8, "현재 가정 손익 상태",
                   f'=IF({op}>=0,"흑자 (BEP 통과)","적자 (BEP 미달)")', "General")

B["B10"] = "민감도 — 배송일당 주문 수별 월 영업이익"
B.merge_cells("B10:I10")
style_block(B, "B10:I10", font=sec_font, fill=sec_fill)

hdr = ["배송일당 주문수", "월 주문건수", "월 매출", "월 공헌이익", "월 고정비", "월 영업이익", "손익"]
for i, h in enumerate(hdr):
    c = B.cell(row=11, column=2 + i, value=h)
    c.font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1D9E75")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
B.row_dimensions[11].height = 30

scenarios = [20, 40, 60, 80, 100, 120, 150, 200]
r = 12
for d in scenarios:
    dcell = f"$B${r}"
    B.cell(row=r, column=2, value=d).number_format = WON_PLAIN
    # 월 주문건수 = 배송일당 × 월배송일수
    B.cell(row=r, column=3, value=f"={dcell}*{mdays}").number_format = WON_PLAIN
    # 월 매출
    B.cell(row=r, column=4, value=f"=$C{r}*{ref['aov']}").number_format = WON
    # 월 공헌이익 = 건당 공헌이익 × 월주문건수
    B.cell(row=r, column=5, value=f"=$C{r}*{cm_unit}").number_format = WON
    # 월 고정비
    B.cell(row=r, column=6, value=f"={f_sum}").number_format = WON
    # 월 영업이익
    B.cell(row=r, column=7, value=f"=$E{r}-$F{r}").number_format = WON
    # 손익
    B.cell(row=r, column=8,
           value=f'=IF($G{r}>=0,"흑자","적자")').number_format = "General"
    for col in range(2, 9):
        cell = B.cell(row=r, column=col)
        cell.border = box
        cell.alignment = Alignment(horizontal="right")
        if cell.font is None or cell.font.name != "맑은 고딕":
            cell.font = calc_font
    r += 1

B.cell(row=r + 1, column=2,
       value="※ 노란 셀(가정값 시트)을 바꾸면 모든 수치가 자동 재계산됩니다.").font = note_font
B.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=8)

# 시트 순서: 가정값 → 손익 → 손익분기
wb.move_sheet("손익분기", offset=0)
wb.active = 0

out = r"c:\Users\js\Desktop\dev\w2o-salada\새벽배송_손익분기_시뮬레이터.xlsx"
wb.save(out)
print("저장 완료:", out)
